# -*- coding: utf-8 -*-
"""
엑셀 검색/취합 도구
====================
여러 엑셀 파일(또는 여러 시트)에서 지정한 검색어를, 지정한 열에서 찾아
지정한 열들의 정보를 가져와서, 사용자가 정의한 출력 열 순서/헤더명대로
하나의 새 엑셀 파일로 저장하는 GUI 프로그램.

핵심 특징
---------
1. GUI (tkinter) 기반
2. 검색어(다건, 줄바꿈으로 구분)를 입력하면 지정한 시트의 지정한 열에서 검색
3. 소스(탭) 설정: 탭이름(라벨), 파일, 시트, 검색할 열, 가져올 열들을
   자유롭게 지정 가능. 여러 개를 추가/삭제/사용-제외 가능
4. VLOOKUP과 달리 검색열보다 왼쪽/오른쪽 상관없이 지정한 열의 값을 가져옴
5. 검색어가 여러 행에 중복으로 존재해도 전부 결과에 포함 (누락 없음)
6. 출력 열 설정: 검색어를 표시할 열, 각 소스에서 가져온 값(별칭)을
   출력할 열을 지정하고 헤더명도 자유롭게 지정 가능
7. 결과를 xlsx로 저장 후 자동으로 열기

필요 패키지
-----------
    pip install openpyxl

실행
----
    python excel_search_merger.py
"""

import os
import re
import sys
import json
import platform
import threading
import subprocess
import traceback
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Alignment
except ImportError:
    print("openpyxl 이 설치되어 있지 않습니다. 다음 명령으로 설치하세요:")
    print("    pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 데이터 모델
# ---------------------------------------------------------------------------

SEARCH_TERM_KEY = "__search_term__"
SOURCE_LABEL_KEY = "__source_label__"

SPECIAL_FIELDS = {
    SEARCH_TERM_KEY: "[검색어]",
    SOURCE_LABEL_KEY: "[출처 탭이름]",
}

# 소스(탭) 설정 / 출력 열 설정을 저장할 기본 위치 (홈 디렉토리)
DEFAULT_CONFIG_PATH = os.path.join(os.path.expanduser("~"), "excel_search_merger_settings.json")


@dataclass
class ColumnMap:
    """소스 시트의 원본 열 헤더명 -> 출력에서 쓸 별칭(alias)"""
    source_header: str
    alias: str


@dataclass
class SourceConfig:
    label: str                      # 사용자가 지정하는 탭(소스) 이름
    file_path: str
    sheet_name: str
    search_header: str              # 검색할 열의 헤더명 (또는 헤더 없음일 때는 열 문자, 예: "C")
    match_mode: str = "contains"    # "exact" | "contains"
    enabled: bool = True
    has_header: bool = True         # False면 지정한 행부터 데이터로 취급하고 열을 A,B,C.. 문자로 지정
    start_row: int = 1              # 헤더가 있으면 헤더가 위치한 행 번호, 없으면 데이터가 시작하는 행 번호
                                     # (위쪽에 빈 행/제목行/병합된 행이 있을 때 그 다음 실제 행을 지정)
    resolve_merges: bool = False    # True면 병합된 셀 값을 병합 범위 전체에 채워 넣음(느림). 필요할 때만 켤 것
    columns: List[ColumnMap] = field(default_factory=list)


@dataclass
class OutputColumn:
    field_key: str    # SEARCH_TERM_KEY / SOURCE_LABEL_KEY / alias 문자열
    header: str        # 엑셀에 실제로 찍힐 헤더명


# ---------------------------------------------------------------------------
# 엑셀 읽기 캐시 & 검색 로직
# ---------------------------------------------------------------------------

def _read_sheet_fast(file_path: str, sheet_name: str) -> Tuple[List[list], int]:
    """read_only 모드로 값만 빠르게 읽는다. 병합 셀 보정은 하지 않음.
    read_only 모드는 병합 정보를 제공하지 않는 대신, 셀 객체를 만들지 않고
    스트리밍으로 읽기 때문에 수천~수만 행짜리 파일에서도 훨씬 빠르고
    메모리를 적게 쓴다. 대부분의 파일(병합 셀이 없는 표 형태)에는 이 방식으로 충분하다.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 를 '{os.path.basename(file_path)}' 에서 찾을 수 없습니다.")
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    max_col = max((len(r) for r in rows), default=0)
    for row in rows:
        if len(row) < max_col:
            row.extend([None] * (max_col - len(row)))
    return rows, max_col


def _read_sheet_with_merges(file_path: str, sheet_name: str) -> Tuple[List[list], int]:
    """시트의 모든 행을 읽되, 병합된 셀은 병합 범위 전체에 대표(좌상단) 값을
    채워 넣은 2차원 리스트로 반환한다. (헤더/값이 병합되어 있어 특정 행에서만
    값이 잡히는 문제를 해결하기 위함)

    반환값: (rows, max_col) - rows는 각 행이 list인 2차원 리스트, max_col은 시트의 최대 열 수.
    """
    # 병합 정보(merged_cells)는 read_only 모드에서 제공되지 않으므로 일반 모드로 읽는다.
    wb = load_workbook(file_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 를 '{os.path.basename(file_path)}' 에서 찾을 수 없습니다.")
        ws = wb[sheet_name]

        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        max_col = ws.max_column or (max((len(r) for r in rows), default=0))
        for row in rows:
            if len(row) < max_col:
                row.extend([None] * (max_col - len(row)))

        for rng in ws.merged_cells.ranges:
            anchor_r, anchor_c = rng.min_row, rng.min_col
            if anchor_r - 1 >= len(rows) or anchor_c - 1 >= max_col:
                continue
            anchor_val = rows[anchor_r - 1][anchor_c - 1]
            for r in range(rng.min_row, rng.max_row + 1):
                if r - 1 >= len(rows):
                    continue
                row = rows[r - 1]
                for c in range(rng.min_col, rng.max_col + 1):
                    if c - 1 < len(row) and row[c - 1] is None:
                        row[c - 1] = anchor_val
        return rows, max_col
    finally:
        wb.close()


def _split_headers_and_data(rows: List[list], max_col: int, has_header: bool,
                             start_row: int) -> Tuple[List[str], List[tuple]]:
    """읽어들인 2차원 데이터에서, 지정한 시작 행부터 헤더/데이터를 분리한다."""
    idx = max(start_row - 1, 0)
    if has_header:
        header_row = rows[idx] if idx < len(rows) else []
        headers = [("" if h is None else str(h).strip()) for h in header_row]
        data_rows = [tuple(r) for r in rows[idx + 1:]]
    else:
        data_rows = [tuple(r) for r in rows[idx:]]
        headers = [get_column_letter(i) for i in range(1, max_col + 1)]
    return headers, data_rows


class SheetCache:
    """동일 (파일, 시트)의 원본 데이터를 한 번만 디스크에서 읽고 재사용한다.
    소스 설정 다이얼로그와 검색 실행이 이 캐시를 공유하므로, 같은 파일을
    여러 번 참조/편집해도 디스크 읽기는 최초 1회만 발생한다."""

    def __init__(self):
        # (file_path, sheet_name, resolve_merges) -> (rows, max_col) : 원본 2차원 데이터
        self._raw_cache: Dict[Tuple[str, str, bool], Tuple[List[list], int]] = {}
        # (file_path, sheet_name, resolve_merges, has_header, start_row) -> (headers, data_rows)
        self._cache: Dict[Tuple[str, str, bool, bool, int], Tuple[List[str], List[tuple]]] = {}

    def _get_raw(self, file_path: str, sheet_name: str, resolve_merges: bool) -> Tuple[List[list], int]:
        key = (file_path, sheet_name, resolve_merges)
        if key in self._raw_cache:
            return self._raw_cache[key]
        if resolve_merges:
            rows, max_col = _read_sheet_with_merges(file_path, sheet_name)
        else:
            rows, max_col = _read_sheet_fast(file_path, sheet_name)
        self._raw_cache[key] = (rows, max_col)
        return rows, max_col

    def get(self, file_path: str, sheet_name: str, has_header: bool = True,
            start_row: int = 1, resolve_merges: bool = False) -> Tuple[List[str], List[tuple]]:
        key = (file_path, sheet_name, resolve_merges, has_header, start_row)
        if key in self._cache:
            return self._cache[key]

        rows, max_col = self._get_raw(file_path, sheet_name, resolve_merges)
        headers, data_rows = _split_headers_and_data(rows, max_col, has_header, start_row)

        self._cache[key] = (headers, data_rows)
        return headers, data_rows

    def clear(self):
        self._cache.clear()
        self._raw_cache.clear()

    @staticmethod
    def list_sheets(file_path: str) -> List[str]:
        wb = load_workbook(file_path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()


COLUMN_LETTER_RE = re.compile(r"^[A-Za-z]{1,3}$")


def resolve_column_ref(value: str, headers: List[str]) -> str:
    """사용자가 '검색할 열'이나 '원본 열'에 입력한 값을 실제 헤더값으로 변환한다.

    - 입력값이 현재 헤더 목록에 그대로 존재하면 그 값을 그대로 사용한다.
      (헤더 없음 모드에서는 헤더 자체가 열 문자이므로 그대로 매치됨)
    - 그렇지 않고 입력값이 'A', 'B', 'AC' 같은 열 문자 형태이면, 엑셀의 실제
      열 번호로 환산하여 그 위치의 헤더값으로 바꿔준다. (헤더가 있는 시트에서도
      열 문자로 직접 지정할 수 있게 하기 위함)
    - 둘 다 아니면 입력값을 그대로 돌려준다.
    """
    value = (value or "").strip()
    if not value:
        return value
    if value in headers:
        return value
    if COLUMN_LETTER_RE.match(value):
        try:
            idx = column_index_from_string(value.upper()) - 1
        except ValueError:
            return value
        if 0 <= idx < len(headers):
            return headers[idx]
    return value


def search_in_config(config: SourceConfig, headers: List[str], data_rows: List[tuple],
                      term: str, case_insensitive: bool, ignore_space: bool = False
                      ) -> List[Dict[str, object]]:
    """지정한 검색열에서 term 과 일치(또는 포함)하는 모든 행을 찾아
    가져올 열(alias)들의 값을 담은 dict 목록으로 반환. 중복 행 전부 포함."""
    results: List[Dict[str, object]] = []

    if config.search_header not in headers:
        return results
    search_idx = headers.index(config.search_header)

    alias_to_idx = {}
    for cm in config.columns:
        if cm.source_header in headers:
            alias_to_idx[cm.alias] = headers.index(cm.source_header)

    term_norm = str(term).strip()
    if ignore_space:
        term_norm = re.sub(r"\s+", "", term_norm)
    term_cmp = term_norm.lower() if case_insensitive else term_norm

    for row in data_rows:
        if search_idx >= len(row):
            continue
        cell_val = row[search_idx]
        cell_str = "" if cell_val is None else str(cell_val).strip()
        if ignore_space:
            cell_str = re.sub(r"\s+", "", cell_str)
        cell_cmp = cell_str.lower() if case_insensitive else cell_str

        if config.match_mode == "exact":
            matched = cell_cmp == term_cmp
        else:
            matched = term_cmp in cell_cmp if term_cmp else False

        if matched:
            record: Dict[str, object] = {}
            for alias, idx in alias_to_idx.items():
                v = row[idx] if idx < len(row) else None
                record[alias] = v
            results.append(record)

    return results


def open_file_with_default_app(path: str):
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif system == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass  # 자동 열기가 실패해도 저장 자체는 성공했으므로 무시


# ---------------------------------------------------------------------------
# 소스(탭) 설정 다이얼로그
# ---------------------------------------------------------------------------

class SourceConfigDialog(tk.Toplevel):
    """소스 설정(탭) 추가/편집 창"""

    def __init__(self, master, config: Optional[SourceConfig] = None,
                 default_file: Optional[str] = None, default_sheet: Optional[str] = None,
                 default_has_header: Optional[bool] = None, default_start_row: Optional[int] = None):
        super().__init__(master)
        self.title("소스(탭) 설정")
        self.geometry("760x640")
        self.resizable(False, False)
        self.grab_set()

        self.result: Optional[SourceConfig] = None
        self.available_headers: List[str] = []

        self._editing = config
        self._default_file = default_file
        self._default_sheet = default_sheet
        self._default_has_header = default_has_header
        self._default_start_row = default_start_row

        self._build_ui()

        if config is not None:
            self._load_from_config(config)
        elif self._default_file and self.var_same_file.get():
            self._apply_default_file()

    # -- UI 구성 --------------------------------------------------------
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", **pad)
        frm_top.grid_columnconfigure(1, weight=1)

        r = 0
        ttk.Label(frm_top, text="탭 이름(구분용):").grid(row=r, column=0, sticky="w")
        self.var_label = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.var_label, width=30).grid(row=r, column=1, sticky="w")
        r += 1

        ttk.Label(frm_top, text="엑셀 파일:").grid(row=r, column=0, sticky="w")
        self.var_file = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.var_file, state="readonly").grid(
            row=r, column=1, sticky="we", padx=(0, 4))
        ttk.Button(frm_top, text="파일 선택", command=self._choose_file).grid(row=r, column=2, sticky="e")
        r += 1

        self.var_same_file = tk.BooleanVar(value=bool(self._default_file))
        cb_same_file = ttk.Checkbutton(
            frm_top, text="이전에 등록한 것과 동일한 파일 사용 (탭 이름/검색 조건만 바꿔서 추가)",
            variable=self.var_same_file, command=self._on_toggle_same_file
        )
        if self._editing is None:
            cb_same_file.grid(row=r, column=0, columnspan=3, sticky="w")
            if not self._default_file:
                cb_same_file.state(["disabled"])
        r += 1

        ttk.Label(frm_top, text="시트:").grid(row=r, column=0, sticky="w")
        self.var_sheet = tk.StringVar()
        self.cb_sheet = ttk.Combobox(frm_top, textvariable=self.var_sheet, state="readonly", width=30)
        self.cb_sheet.grid(row=r, column=1, sticky="w")
        self.cb_sheet.bind("<<ComboboxSelected>>", lambda e: self._load_headers())
        r += 1

        self.var_no_header = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            frm_top, text="이 시트는 헤더(제목행)가 없음 (열을 A,B,C.. 로 지정)",
            variable=self.var_no_header, command=self._load_headers
        ).grid(row=r, column=0, columnspan=3, sticky="w")
        r += 1

        ttk.Label(frm_top, text="헤더/데이터 시작 행:").grid(row=r, column=0, sticky="w")
        self.var_start_row = tk.IntVar(value=1)
        self.sb_start_row = ttk.Spinbox(
            frm_top, from_=1, to=100000, textvariable=self.var_start_row, width=8,
            command=self._load_headers
        )
        self.sb_start_row.grid(row=r, column=1, sticky="w")
        self.sb_start_row.bind("<FocusOut>", lambda e: self._load_headers())
        self.sb_start_row.bind("<Return>", lambda e: self._load_headers())
        r += 1
        ttk.Label(
            frm_top,
            text="(위쪽에 빈 행/제목行/병합된 행이 있으면 실제 헤더(또는 데이터)가 시작되는 행 번호를 입력. 병합된 셀은 값이 자동으로 채워집니다)",
            foreground="#666", wraplength=680, justify="left"
        ).grid(row=r, column=0, columnspan=3, sticky="w")
        r += 1

        ttk.Label(frm_top, text="검색할 열:").grid(row=r, column=0, sticky="w")
        self.var_search_col = tk.StringVar()
        self.cb_search_col = ttk.Combobox(frm_top, textvariable=self.var_search_col, width=30)
        self.cb_search_col.grid(row=r, column=1, sticky="w")
        r += 1
        ttk.Label(
            frm_top, text="(목록에서 선택하거나, 열 문자를 직접 입력해도 됩니다. 예: C)",
            foreground="#666"
        ).grid(row=r, column=0, columnspan=3, sticky="w")
        r += 1

        ttk.Label(frm_top, text="검색 방식:").grid(row=r, column=0, sticky="w")
        self.var_match = tk.StringVar(value="contains")
        frm_match = ttk.Frame(frm_top)
        frm_match.grid(row=r, column=1, sticky="w")
        ttk.Radiobutton(frm_match, text="부분일치(포함)", variable=self.var_match, value="contains").pack(side="left")
        ttk.Radiobutton(frm_match, text="완전일치", variable=self.var_match, value="exact").pack(side="left")
        r += 1

        # 가져올 열 매핑 영역
        frm_cols = ttk.LabelFrame(self, text="가져올 열 (원본 열 -> 출력 별칭)")
        frm_cols.pack(fill="both", expand=True, **pad)

        self.tree_cols = ttk.Treeview(frm_cols, columns=("source", "alias"), show="headings", height=10)
        self.tree_cols.heading("source", text="원본 열(헤더명)")
        self.tree_cols.heading("alias", text="출력 별칭")
        self.tree_cols.column("source", width=250)
        self.tree_cols.column("alias", width=250)
        self.tree_cols.pack(fill="both", expand=True, padx=6, pady=4)

        frm_add = ttk.Frame(frm_cols)
        frm_add.pack(fill="x", padx=6, pady=4)

        ttk.Label(frm_add, text="원본 열:").pack(side="left")
        self.var_new_source = tk.StringVar()
        self.cb_new_source = ttk.Combobox(frm_add, textvariable=self.var_new_source, width=22)
        self.cb_new_source.pack(side="left", padx=4)
        self.cb_new_source.bind("<<ComboboxSelected>>", self._sync_alias_default)

        ttk.Label(frm_add, text="별칭:").pack(side="left")
        self.var_new_alias = tk.StringVar()
        ttk.Entry(frm_add, textvariable=self.var_new_alias, width=20).pack(side="left", padx=4)

        ttk.Button(frm_add, text="목록에 추가", command=self._add_column_map).pack(side="left", padx=6)
        ttk.Button(frm_add, text="선택 삭제", command=self._remove_column_map).pack(side="left")

        ttk.Label(
            frm_cols,
            text="※ '원본 열'도 목록에서 고르거나, 열 문자(A, B, C..)를 직접 입력할 수 있습니다.",
            foreground="#555"
        ).pack(anchor="w", padx=6)

        # 하단 저장/취소
        frm_bottom = ttk.Frame(self)
        frm_bottom.pack(fill="x", **pad)
        ttk.Button(frm_bottom, text="저장", command=self._on_save).pack(side="right", padx=4)
        ttk.Button(frm_bottom, text="취소", command=self.destroy).pack(side="right")

    # -- 동작 --------------------------------------------------------
    def _choose_file(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        # 사용자가 직접 다른 파일을 골랐으므로 '동일 파일 사용' 체크는 해제
        if hasattr(self, "var_same_file"):
            self.var_same_file.set(False)
        self.var_file.set(path)
        try:
            sheets = SheetCache.list_sheets(path)
        except Exception as e:
            messagebox.showerror("오류", f"파일을 여는 중 오류가 발생했습니다:\n{e}", parent=self)
            return
        self.cb_sheet["values"] = sheets
        if sheets:
            self.var_sheet.set(sheets[0])
            self._load_headers()

    def _apply_default_file(self):
        """이전에 등록한 소스와 동일한 파일/시트/헤더 설정을 그대로 불러온다."""
        path = self._default_file
        if not path:
            return
        self.var_file.set(path)
        try:
            sheets = SheetCache.list_sheets(path)
        except Exception as e:
            messagebox.showerror("오류", f"파일을 여는 중 오류가 발생했습니다:\n{e}", parent=self)
            return
        self.cb_sheet["values"] = sheets
        sheet = self._default_sheet if self._default_sheet in sheets else (sheets[0] if sheets else "")
        self.var_sheet.set(sheet)
        if self._default_has_header is not None:
            self.var_no_header.set(not self._default_has_header)
        if self._default_start_row is not None:
            self.var_start_row.set(self._default_start_row)
        if sheet:
            self._load_headers()

    def _on_toggle_same_file(self):
        if self.var_same_file.get():
            self._apply_default_file()
        else:
            self.var_file.set("")
            self.var_sheet.set("")
            self.cb_sheet["values"] = []
            self.var_search_col.set("")
            self.available_headers = []
            self.cb_search_col["values"] = []
            self.cb_new_source["values"] = []

    def _load_headers(self):
        path = self.var_file.get()
        sheet = self.var_sheet.get()
        if not path or not sheet:
            return
        try:
            start_row = max(int(self.var_start_row.get() or 1), 1)
        except (tk.TclError, ValueError):
            start_row = 1
        try:
            rows, max_col = _read_sheet_with_merges(path, sheet)
            has_header = not self.var_no_header.get()
            headers, _ = _split_headers_and_data(rows, max_col, has_header, start_row)
            if has_header:
                headers = [h for h in headers if h != ""]
        except Exception as e:
            messagebox.showerror("오류", f"시트를 읽는 중 오류가 발생했습니다:\n{e}", parent=self)
            return

        self.available_headers = headers

        self.cb_search_col["values"] = headers
        self.cb_new_source["values"] = headers
        if headers and not self.var_search_col.get():
            self.var_search_col.set(headers[0])

    def _sync_alias_default(self, event=None):
        if not self.var_new_alias.get():
            self.var_new_alias.set(self.var_new_source.get())

    def _add_column_map(self):
        src = self.var_new_source.get().strip()
        alias = self.var_new_alias.get().strip()
        if not src:
            messagebox.showwarning("확인", "원본 열을 선택하세요.", parent=self)
            return
        src = resolve_column_ref(src, self.available_headers)
        if not alias:
            alias = src
        self.tree_cols.insert("", "end", values=(src, alias))
        self.var_new_source.set("")
        self.var_new_alias.set("")

    def _remove_column_map(self):
        for item in self.tree_cols.selection():
            self.tree_cols.delete(item)

    def _load_from_config(self, config: SourceConfig):
        self.var_label.set(config.label)
        self.var_file.set(config.file_path)
        try:
            sheets = SheetCache.list_sheets(config.file_path)
            self.cb_sheet["values"] = sheets
        except Exception:
            pass
        self.var_sheet.set(config.sheet_name)
        self.var_no_header.set(not config.has_header)
        self.var_start_row.set(config.start_row)
        self._load_headers()
        self.var_search_col.set(config.search_header)
        self.var_match.set(config.match_mode)
        for cm in config.columns:
            self.tree_cols.insert("", "end", values=(cm.source_header, cm.alias))

    def _on_save(self):
        label = self.var_label.get().strip()
        file_path = self.var_file.get().strip()
        sheet = self.var_sheet.get().strip()
        search_col = self.var_search_col.get().strip()

        if not label:
            messagebox.showwarning("확인", "탭 이름을 입력하세요.", parent=self)
            return
        if not file_path or not os.path.isfile(file_path):
            messagebox.showwarning("확인", "유효한 엑셀 파일을 선택하세요.", parent=self)
            return
        if not sheet:
            messagebox.showwarning("확인", "시트를 선택하세요.", parent=self)
            return
        if not search_col:
            messagebox.showwarning("확인", "검색할 열을 선택하세요.", parent=self)
            return
        search_col = resolve_column_ref(search_col, self.available_headers)

        columns = []
        for item in self.tree_cols.get_children():
            src, alias = self.tree_cols.item(item, "values")
            columns.append(ColumnMap(source_header=src, alias=alias))
        if not columns:
            messagebox.showwarning("확인", "가져올 열을 최소 1개 이상 추가하세요.", parent=self)
            return

        enabled = self._editing.enabled if self._editing is not None else True

        try:
            start_row = max(int(self.var_start_row.get() or 1), 1)
        except (tk.TclError, ValueError):
            start_row = 1

        self.result = SourceConfig(
            label=label,
            file_path=file_path,
            sheet_name=sheet,
            search_header=search_col,
            match_mode=self.var_match.get(),
            enabled=enabled,
            has_header=not self.var_no_header.get(),
            start_row=start_row,
            columns=columns,
        )
        self.destroy()


# ---------------------------------------------------------------------------
# 메인 애플리케이션
# ---------------------------------------------------------------------------

class ExcelSearchMergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("엑셀 검색/취합 도구")
        self.geometry("900x680")

        self.configs: List[SourceConfig] = []
        self.output_columns: List[OutputColumn] = []
        self.cache = SheetCache()
        self.last_results: List[Dict[str, object]] = []

        self._build_ui()
        self._load_config(silent=True)  # 이전에 저장한 상태가 있으면 자동으로 불러오기

    # ------------------------------------------------------------------
    def _build_ui(self):
        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=6, pady=(6, 0))

        ttk.Button(toolbar, text="상태 저장", command=self._save_config).pack(side="left", padx=2)
        ttk.Button(toolbar, text="상태 불러오기", command=lambda: self._load_config(silent=False)).pack(
            side="left", padx=2)
        ttk.Button(toolbar, text="초기화", command=self._reset_all).pack(side="left", padx=2)
        ttk.Label(toolbar, text=f"저장 위치: {DEFAULT_CONFIG_PATH}", foreground="#666").pack(side="left", padx=10)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.tab_sources = ttk.Frame(nb)
        self.tab_output = ttk.Frame(nb)
        self.tab_run = ttk.Frame(nb)

        nb.add(self.tab_sources, text="① 소스(탭) 설정")
        nb.add(self.tab_output, text="② 출력 열 설정")
        nb.add(self.tab_run, text="③ 검색 실행 / 결과")

        self._build_tab_sources()
        self._build_tab_output()
        self._build_tab_run()

    # -- 탭 1: 소스 설정 --------------------------------------------------
    def _build_tab_sources(self):
        frm = self.tab_sources

        self.tree_sources = ttk.Treeview(
            frm, columns=("use", "label", "file", "sheet", "header", "start_row", "search", "cols"),
            show="headings", height=14
        )
        headers = [
            ("use", "사용", 50), ("label", "탭이름", 120), ("file", "파일", 200),
            ("sheet", "시트", 90), ("header", "헤더", 60), ("start_row", "시작행", 60),
            ("search", "검색열", 90), ("cols", "가져올열 수", 90),
        ]
        for key, text, w in headers:
            self.tree_sources.heading(key, text=text)
            self.tree_sources.column(key, width=w, anchor="center" if key in ("use", "cols", "start_row") else "w")
        self.tree_sources.pack(fill="both", expand=True, padx=8, pady=8)

        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill="x", padx=8, pady=4)
        ttk.Button(btn_frm, text="추가", command=self._add_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="편집", command=self._edit_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="삭제", command=self._delete_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="사용/제외 전환", command=self._toggle_source).pack(side="left", padx=2)

    def _refresh_sources_tree(self):
        self.tree_sources.delete(*self.tree_sources.get_children())
        for i, c in enumerate(self.configs):
            self.tree_sources.insert(
                "", "end", iid=str(i),
                values=("O" if c.enabled else "X", c.label, os.path.basename(c.file_path),
                        c.sheet_name, "있음" if c.has_header else "없음", c.start_row,
                        c.search_header, len(c.columns))
            )
        self._refresh_output_field_choices()

    def _add_source(self):
        last = self.configs[-1] if self.configs else None
        dlg = SourceConfigDialog(
            self,
            default_file=last.file_path if last else None,
            default_sheet=last.sheet_name if last else None,
            default_has_header=last.has_header if last else None,
            default_start_row=last.start_row if last else None,
        )
        self.wait_window(dlg)
        if dlg.result:
            self.configs.append(dlg.result)
            self._refresh_sources_tree()

    def _edit_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            messagebox.showinfo("확인", "편집할 항목을 선택하세요.")
            return
        idx = int(sel[0])
        dlg = SourceConfigDialog(self, config=self.configs[idx])
        self.wait_window(dlg)
        if dlg.result:
            self.configs[idx] = dlg.result
            self._refresh_sources_tree()

    def _delete_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            return
        idx = int(sel[0])
        del self.configs[idx]
        self._refresh_sources_tree()

    def _toggle_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            return
        idx = int(sel[0])
        self.configs[idx].enabled = not self.configs[idx].enabled
        self._refresh_sources_tree()

    # -- 탭 2: 출력 열 설정 -------------------------------------------------
    def _build_tab_output(self):
        frm = self.tab_output

        self.tree_output = ttk.Treeview(frm, columns=("header", "field"), show="headings", height=12)
        self.tree_output.heading("header", text="출력 헤더명")
        self.tree_output.heading("field", text="필드(소스)")
        self.tree_output.column("header", width=250)
        self.tree_output.column("field", width=250)
        self.tree_output.pack(fill="both", expand=True, padx=8, pady=8)

        frm_add = ttk.Frame(frm)
        frm_add.pack(fill="x", padx=8, pady=4)

        ttk.Label(frm_add, text="필드:").pack(side="left")
        self.var_out_field = tk.StringVar()
        self.cb_out_field = ttk.Combobox(frm_add, textvariable=self.var_out_field, state="readonly", width=28)
        self.cb_out_field.pack(side="left", padx=4)

        ttk.Label(frm_add, text="헤더명:").pack(side="left")
        self.var_out_header = tk.StringVar()
        ttk.Entry(frm_add, textvariable=self.var_out_header, width=22).pack(side="left", padx=4)

        self.var_out_add_btn = tk.StringVar(value="추가")
        self._editing_output_idx: Optional[int] = None
        ttk.Button(frm_add, textvariable=self.var_out_add_btn, command=self._add_output_column).pack(
            side="left", padx=6)
        ttk.Button(frm_add, text="선택 수정", command=self._edit_output_column).pack(side="left", padx=2)
        ttk.Button(frm_add, text="수정 취소", command=self._cancel_edit_output_column).pack(side="left", padx=2)
        ttk.Button(frm_add, text="선택 삭제", command=self._remove_output_column).pack(side="left", padx=2)
        ttk.Button(frm_add, text="위로", command=lambda: self._move_output_column(-1)).pack(side="left", padx=2)
        ttk.Button(frm_add, text="아래로", command=lambda: self._move_output_column(1)).pack(side="left", padx=2)

        self.tree_output.bind("<Double-1>", lambda e: self._edit_output_column())

        ttk.Label(
            frm,
            text="※ 필드는 [검색어], [출처 탭이름], 또는 소스 설정에서 지정한 '출력 별칭' 중에서 고를 수 있습니다.\n"
                 "※ 목록에서 더블클릭하거나 '선택 수정'을 누르면 값을 불러와 바로 수정할 수 있습니다.",
            foreground="#555", justify="left"
        ).pack(anchor="w", padx=8, pady=(0, 8))

    def _refresh_output_field_choices(self):
        aliases = []
        for c in self.configs:
            for cm in c.columns:
                if cm.alias not in aliases:
                    aliases.append(cm.alias)
        choices = list(SPECIAL_FIELDS.values()) + aliases
        self.cb_out_field["values"] = choices

    def _field_label_to_key(self, label: str) -> str:
        for k, v in SPECIAL_FIELDS.items():
            if v == label:
                return k
        return label  # alias 그대로 키로 사용

    def _field_key_to_label(self, key: str) -> str:
        return SPECIAL_FIELDS.get(key, key)

    def _refresh_output_tree(self):
        self.tree_output.delete(*self.tree_output.get_children())
        for oc in self.output_columns:
            self.tree_output.insert("", "end", values=(oc.header, self._field_key_to_label(oc.field_key)))

    def _add_output_column(self):
        field_label = self.var_out_field.get().strip()
        header = self.var_out_header.get().strip()
        if not field_label:
            messagebox.showwarning("확인", "필드를 선택하세요.")
            return
        if not header:
            header = field_label
        key = self._field_label_to_key(field_label)

        if self._editing_output_idx is not None:
            self.output_columns[self._editing_output_idx] = OutputColumn(field_key=key, header=header)
        else:
            self.output_columns.append(OutputColumn(field_key=key, header=header))

        self._cancel_edit_output_column()
        self._refresh_output_tree()

    def _edit_output_column(self):
        sel = self.tree_output.selection()
        if not sel:
            messagebox.showinfo("확인", "수정할 항목을 선택하세요.")
            return
        idx = self.tree_output.index(sel[0])
        oc = self.output_columns[idx]
        self.var_out_field.set(self._field_key_to_label(oc.field_key))
        self.var_out_header.set(oc.header)
        self._editing_output_idx = idx
        self.var_out_add_btn.set("수정 저장")

    def _cancel_edit_output_column(self):
        self._editing_output_idx = None
        self.var_out_add_btn.set("추가")
        self.var_out_field.set("")
        self.var_out_header.set("")

    def _remove_output_column(self):
        sel = self.tree_output.selection()
        if not sel:
            return
        indices = sorted((self.tree_output.index(i) for i in sel), reverse=True)
        for i in indices:
            del self.output_columns[i]
        self._cancel_edit_output_column()
        self._refresh_output_tree()

    def _move_output_column(self, delta: int):
        sel = self.tree_output.selection()
        if not sel:
            return
        self._cancel_edit_output_column()
        idx = self.tree_output.index(sel[0])
        new_idx = idx + delta
        if 0 <= new_idx < len(self.output_columns):
            self.output_columns[idx], self.output_columns[new_idx] = (
                self.output_columns[new_idx], self.output_columns[idx]
            )
            self._refresh_output_tree()
            self.tree_output.selection_set(str(new_idx))

    # -- 탭 3: 검색 실행 / 결과 ---------------------------------------------
    def _build_tab_run(self):
        frm = self.tab_run

        ttk.Label(frm, text="검색어 입력 (한 줄에 하나씩):").pack(anchor="w", padx=8, pady=(8, 0))
        self.txt_terms = tk.Text(frm, height=6)
        self.txt_terms.pack(fill="x", padx=8, pady=4)

        opt_frm = ttk.Frame(frm)
        opt_frm.pack(fill="x", padx=8)
        self.var_case_insensitive = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_frm, text="대소문자 구분 안 함", variable=self.var_case_insensitive).pack(side="left")

        self.var_ignore_space = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frm, text="띄어쓰기 차이 무시 (예: '홍 길동' = '홍길동')",
            variable=self.var_ignore_space
        ).pack(side="left", padx=16)

        self.var_merge_mode = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frm, text="여러 탭 결과를 검색어 기준으로 한 행에 병합(권장)",
            variable=self.var_merge_mode
        ).pack(side="left", padx=16)

        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill="x", padx=8, pady=6)
        ttk.Button(btn_frm, text="검색 실행", command=self._run_search).pack(side="left")
        ttk.Button(btn_frm, text="엑셀로 저장", command=self._save_results).pack(side="left", padx=6)

        self.lbl_status = ttk.Label(frm, text="검색 결과: 0건")
        self.lbl_status.pack(anchor="w", padx=8)

        self.tree_result = ttk.Treeview(frm, show="headings", height=16)
        self.tree_result.pack(fill="both", expand=True, padx=8, pady=8)

    def _run_search(self):
        if not self.configs:
            messagebox.showwarning("확인", "소스(탭) 설정을 먼저 추가하세요.")
            return
        if not self.output_columns:
            messagebox.showwarning("확인", "출력 열 설정을 먼저 추가하세요.")
            return

        terms = [t.strip() for t in self.txt_terms.get("1.0", "end").splitlines() if t.strip()]
        if not terms:
            messagebox.showwarning("확인", "검색어를 한 개 이상 입력하세요.")
            return

        self.cache.clear()
        results: List[Dict[str, object]] = []
        case_ins = self.var_case_insensitive.get()
        ignore_space = self.var_ignore_space.get()
        merge_mode = self.var_merge_mode.get()

        try:
            active_configs = [c for c in self.configs if c.enabled]
            for term in terms:
                if merge_mode:
                    # 병합 모드: 검색어 하나당, 활성화된 탭들의 매치 결과를 한 데이터셋으로 묶는다.
                    # - 어떤 탭에서 매치가 여러 건이면(예: 구매내역) 그 건수만큼 행을 만든다.
                    # - 매치가 1건뿐인 탭(예: 직원목록)의 값은 모든 행에 동일하게 채워 넣는다(반복).
                    per_config_matches = []
                    for cfg in active_configs:
                        headers, rows = self.cache.get(cfg.file_path, cfg.sheet_name, cfg.has_header, cfg.start_row)
                        matches = search_in_config(cfg, headers, rows, term, case_ins, ignore_space)
                        per_config_matches.append((cfg, matches))

                    match_counts = [len(m) for _, m in per_config_matches if len(m) > 0]
                    if not match_counts:
                        continue  # 이 검색어는 어떤 탭에서도 매치되지 않음
                    max_count = max(match_counts)

                    for i in range(max_count):
                        rec: Dict[str, object] = {SEARCH_TERM_KEY: term}
                        labels = []
                        for cfg, matches in per_config_matches:
                            if not matches:
                                continue
                            labels.append(cfg.label)
                            if len(matches) == 1:
                                rec.update(matches[0])       # 1건이면 모든 행에 반복 채움
                            elif i < len(matches):
                                rec.update(matches[i])       # 여러 건이면 순서대로 배치
                            # i가 해당 탭의 매치 건수를 넘어서면 그 탭의 값은 빈칸으로 둠
                        rec[SOURCE_LABEL_KEY] = "+".join(labels)
                        results.append(rec)
                else:
                    # 쌓기 모드(기존 방식): 탭별로 매치된 행을 각각 별도의 결과 행으로 추가
                    for cfg in active_configs:
                        headers, rows = self.cache.get(cfg.file_path, cfg.sheet_name, cfg.has_header, cfg.start_row)
                        matches = search_in_config(cfg, headers, rows, term, case_ins, ignore_space)
                        for m in matches:
                            rec = dict(m)
                            rec[SEARCH_TERM_KEY] = term
                            rec[SOURCE_LABEL_KEY] = cfg.label
                            results.append(rec)
        except Exception as e:
            messagebox.showerror("오류", f"검색 중 오류가 발생했습니다:\n{e}\n\n{traceback.format_exc()}")
            return

        self.last_results = results
        self._show_results(results)

        if not results:
            diag_lines = ["검색 결과가 없습니다. 아래 진단 정보를 확인해 주세요.\n"]
            for cfg in active_configs:
                try:
                    headers, rows = self.cache.get(cfg.file_path, cfg.sheet_name, cfg.has_header, cfg.start_row)
                except Exception as e:
                    diag_lines.append(f"[{cfg.label}] 파일/시트를 읽는 중 오류: {e}")
                    continue

                if cfg.search_header not in headers:
                    diag_lines.append(
                        f"[{cfg.label}] 설정된 검색열 '{cfg.search_header}' 을(를) "
                        f"시트의 지정된 헤더 행(현재 {cfg.start_row}행)에서 찾지 못했습니다.\n"
                        f"    이 시트의 실제 헤더 목록: {headers}\n"
                        f"    → 시트 맨 위에 제목/병합된 셀이 있어서 지정한 행이 실제 헤더가 아닐 수 있습니다.\n"
                        f"    → 소스 설정에서 '헤더/데이터 시작 행' 값을 실제 헤더가 있는 행 번호로 조정하세요.\n"
                        f"    → 이 시트에 헤더(제목행) 자체가 없다면, '헤더 없음' 옵션을 켜고 "
                        f"검색열을 A,B,C.. 열 문자로 지정하세요."
                    )
                    continue

                total = 0
                for term in terms:
                    total += len(search_in_config(cfg, headers, rows, term, case_ins, ignore_space))

                search_idx = headers.index(cfg.search_header)
                sample_vals = []
                seen = set()
                for row in rows:
                    if search_idx < len(row):
                        v = row[search_idx]
                        s = "" if v is None else str(v)
                        if s not in seen:
                            seen.add(s)
                            sample_vals.append(repr(s))
                    if len(sample_vals) >= 10:
                        break

                diag_lines.append(
                    f"[{cfg.label}] 검색열은 정상 인식됨(데이터 {len(rows)}행). "
                    f"입력한 검색어 전체와의 매치 수: {total}건\n"
                    f"    현재 매칭 방식: {'완전일치' if cfg.match_mode == 'exact' else '부분일치'}\n"
                    f"    입력한 검색어: {terms}\n"
                    f"    검색열의 실제 값 샘플(최대 10개, 형식 비교용): {sample_vals}"
                )
            messagebox.showinfo("검색 결과 없음 - 진단 정보", "\n\n".join(diag_lines))

    def _show_results(self, results: List[Dict[str, object]]):
        cols = [oc.header for oc in self.output_columns]
        self.tree_result["columns"] = cols
        self.tree_result.delete(*self.tree_result.get_children())
        for c in cols:
            self.tree_result.heading(c, text=c, anchor="center")
            self.tree_result.column(c, width=140, anchor="center")

        for rec in results:
            row_vals = []
            for oc in self.output_columns:
                v = rec.get(oc.field_key, "")
                row_vals.append("" if v is None else v)
            self.tree_result.insert("", "end", values=row_vals)

        self.lbl_status.config(text=f"검색 결과: {len(results)}건")

    def _save_results(self):
        if not self.last_results:
            messagebox.showinfo("확인", "먼저 검색을 실행하세요.")
            return

        path = filedialog.asksaveasfilename(
            title="결과 저장", defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")]
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "결과"

            headers = [oc.header for oc in self.output_columns]
            ws.append(headers)

            for rec in self.last_results:
                row = []
                for oc in self.output_columns:
                    v = rec.get(oc.field_key, "")
                    row.append("" if v is None else v)
                ws.append(row)

            # 대략적인 열 너비 자동 조정
            for col_idx, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for rec in self.last_results:
                    v = rec.get(self.output_columns[col_idx - 1].field_key, "")
                    max_len = max(max_len, len(str(v)))
                letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[letter].width = min(max_len + 4, 60)

            # 헤더/데이터 모든 셀 가운데 정렬
            center_align = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = center_align

            wb.save(path)
        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류가 발생했습니다:\n{e}")
            return

        messagebox.showinfo("완료", f"저장이 완료되었습니다:\n{path}")
        open_file_with_default_app(path)

    # -- 상태 저장 / 불러오기 / 초기화 -------------------------------------
    def _config_to_dict(self) -> dict:
        return {
            "sources": [
                {
                    "label": c.label,
                    "file_path": c.file_path,
                    "sheet_name": c.sheet_name,
                    "search_header": c.search_header,
                    "match_mode": c.match_mode,
                    "enabled": c.enabled,
                    "has_header": c.has_header,
                    "start_row": c.start_row,
                    "columns": [{"source_header": cm.source_header, "alias": cm.alias} for cm in c.columns],
                }
                for c in self.configs
            ],
            "output_columns": [
                {"field_key": oc.field_key, "header": oc.header} for oc in self.output_columns
            ],
        }

    def _apply_config_dict(self, data: dict):
        configs = []
        for cd in data.get("sources", []):
            cols = [
                ColumnMap(source_header=cm.get("source_header", ""), alias=cm.get("alias", ""))
                for cm in cd.get("columns", [])
            ]
            configs.append(SourceConfig(
                label=cd.get("label", ""),
                file_path=cd.get("file_path", ""),
                sheet_name=cd.get("sheet_name", ""),
                search_header=cd.get("search_header", ""),
                match_mode=cd.get("match_mode", "contains"),
                enabled=cd.get("enabled", True),
                has_header=cd.get("has_header", True),
                start_row=cd.get("start_row", 1),
                columns=cols,
            ))
        output_columns = [
            OutputColumn(field_key=ocd.get("field_key", ""), header=ocd.get("header", ""))
            for ocd in data.get("output_columns", [])
        ]

        self.configs = configs
        self.output_columns = output_columns
        self._refresh_sources_tree()
        self._refresh_output_tree()

    def _save_config(self):
        data = self._config_to_dict()
        try:
            with open(DEFAULT_CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("오류", f"상태 저장 중 오류가 발생했습니다:\n{e}")
            return
        messagebox.showinfo("완료", f"현재 소스/출력 설정을 저장했습니다:\n{DEFAULT_CONFIG_PATH}")

    def _load_config(self, silent: bool = False):
        if not os.path.isfile(DEFAULT_CONFIG_PATH):
            if not silent:
                messagebox.showinfo("확인", "저장된 상태 파일이 없습니다.")
            return
        try:
            with open(DEFAULT_CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            self._apply_config_dict(data)
        except Exception as e:
            if not silent:
                messagebox.showerror("오류", f"상태 불러오기 중 오류가 발생했습니다:\n{e}")
            return
        if not silent:
            messagebox.showinfo("완료", "저장된 소스/출력 설정을 불러왔습니다.")

    def _reset_all(self):
        if not messagebox.askyesno("확인", "소스 설정, 출력 열 설정, 검색어, 결과를 모두 초기화하시겠습니까?\n"
                                          "(저장된 상태 파일은 삭제되지 않습니다.)"):
            return
        self.configs = []
        self.output_columns = []
        self.last_results = []
        self._refresh_sources_tree()
        self._refresh_output_tree()
        self.txt_terms.delete("1.0", "end")
        self.tree_result["columns"] = ()
        self.tree_result.delete(*self.tree_result.get_children())
        self.lbl_status.config(text="검색 결과: 0건")


def main():
    app = ExcelSearchMergerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
