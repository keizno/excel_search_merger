# -*- coding: utf-8 -*-
"""
엑셀 검색/취합 도구
====================
여러 엑셀 파일(또는 여러 시트)에서 지정한 검색어를, 지정한 열에서 찾아
지정한 열들의 정보를 가져와서, 사용자가 정의한 출력 열 순서/헤더명대로
하나의 새 엑셀 파일로 저장하는 GUI 프로그램.

핵심 특징
---------
1. GUI (tkinter) 기반
2. 검색어(다건, 줄바꿈으로 구분)를 입력하면 지정한 시트의 지정한 열에서 검색
3. 소스(탭) 설정: 탭이름(라벨), 파일, 시트, 검색할 열, 가져올 열들을
   자유롭게 지정 가능. 여러 개를 추가/삭제/사용-제외 가능
4. VLOOKUP과 달리 검색열보다 왼쪽/오른쪽 상관없이 지정한 열의 값을 가져옴
5. 검색어가 여러 행에 중복으로 존재해도 전부 결과에 포함 (누락 없음)
6. 출력 열 설정: 검색어를 표시할 열, 각 소스에서 가져온 값(별칭)을
   출력할 열을 지정하고 헤더명도 자유롭게 지정 가능
7. 결과를 xlsx로 저장 후 자동으로 열기

필요 패키지
-----------
    pip install openpyxl

실행
----
    python excel_search_merger.py
"""

import os
import sys
import json
import platform
import subprocess
import traceback
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl 이 설치되어 있지 않습니다. 다음 명령으로 설치하세요:")
    print("    pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 데이터 모델
# ---------------------------------------------------------------------------

SEARCH_TERM_KEY = "__search_term__"
SOURCE_LABEL_KEY = "__source_label__"

SPECIAL_FIELDS = {
    SEARCH_TERM_KEY: "[검색어]",
    SOURCE_LABEL_KEY: "[출처 탭이름]",
}

# 소스(탭) 설정 / 출력 열 설정을 저장할 기본 위치 (홈 디렉토리)
DEFAULT_CONFIG_PATH = os.path.join(os.path.expanduser("~"), "excel_search_merger_settings.json")


@dataclass
class ColumnMap:
    """소스 시트의 원본 열 헤더명 -> 출력에서 쓸 별칭(alias)"""
    source_header: str
    alias: str


@dataclass
class SourceConfig:
    label: str                      # 사용자가 지정하는 탭(소스) 이름
    file_path: str
    sheet_name: str
    search_header: str              # 검색할 열의 헤더명
    match_mode: str = "contains"    # "exact" | "contains"
    enabled: bool = True
    columns: List[ColumnMap] = field(default_factory=list)


@dataclass
class OutputColumn:
    field_key: str    # SEARCH_TERM_KEY / SOURCE_LABEL_KEY / alias 문자열
    header: str        # 엑셀에 실제로 찍힐 헤더명


# ---------------------------------------------------------------------------
# 엑셀 읽기 캐시 & 검색 로직
# ---------------------------------------------------------------------------

class SheetCache:
    """동일 (파일, 시트)를 여러 검색어에 대해 반복해서 여는 것을 방지"""

    def __init__(self):
        self._cache: Dict[Tuple[str, str], Tuple[List[str], List[tuple]]] = {}

    def get(self, file_path: str, sheet_name: str) -> Tuple[List[str], List[tuple]]:
        key = (file_path, sheet_name)
        if key in self._cache:
            return self._cache[key]

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"시트 '{sheet_name}' 를 '{os.path.basename(file_path)}' 에서 찾을 수 없습니다.")
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                header_row = ()
            headers = [("" if h is None else str(h).strip()) for h in header_row]
            data_rows = list(rows_iter)
        finally:
            wb.close()

        self._cache[key] = (headers, data_rows)
        return headers, data_rows

    def clear(self):
        self._cache.clear()

    @staticmethod
    def list_sheets(file_path: str) -> List[str]:
        wb = load_workbook(file_path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()


def search_in_config(config: SourceConfig, headers: List[str], data_rows: List[tuple],
                      term: str, case_insensitive: bool) -> List[Dict[str, object]]:
    """지정한 검색열에서 term 과 일치(또는 포함)하는 모든 행을 찾아
    가져올 열(alias)들의 값을 담은 dict 목록으로 반환. 중복 행 전부 포함."""
    results: List[Dict[str, object]] = []

    if config.search_header not in headers:
        return results
    search_idx = headers.index(config.search_header)

    alias_to_idx = {}
    for cm in config.columns:
        if cm.source_header in headers:
            alias_to_idx[cm.alias] = headers.index(cm.source_header)

    term_norm = str(term).strip()
    term_cmp = term_norm.lower() if case_insensitive else term_norm

    for row in data_rows:
        if search_idx >= len(row):
            continue
        cell_val = row[search_idx]
        cell_str = "" if cell_val is None else str(cell_val).strip()
        cell_cmp = cell_str.lower() if case_insensitive else cell_str

        if config.match_mode == "exact":
            matched = cell_cmp == term_cmp
        else:
            matched = term_cmp in cell_cmp if term_cmp else False

        if matched:
            record: Dict[str, object] = {}
            for alias, idx in alias_to_idx.items():
                v = row[idx] if idx < len(row) else None
                record[alias] = v
            results.append(record)

    return results


def open_file_with_default_app(path: str):
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif system == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass  # 자동 열기가 실패해도 저장 자체는 성공했으므로 무시


# ---------------------------------------------------------------------------
# 소스(탭) 설정 다이얼로그
# ---------------------------------------------------------------------------

class SourceConfigDialog(tk.Toplevel):
    """소스 설정(탭) 추가/편집 창"""

    def __init__(self, master, config: Optional[SourceConfig] = None):
        super().__init__(master)
        self.title("소스(탭) 설정")
        self.geometry("620x560")
        self.resizable(False, False)
        self.grab_set()

        self.result: Optional[SourceConfig] = None
        self.available_headers: List[str] = []

        self._editing = config
        self._build_ui()

        if config is not None:
            self._load_from_config(config)

    # -- UI 구성 --------------------------------------------------------
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", **pad)

        ttk.Label(frm_top, text="탭 이름(구분용):").grid(row=0, column=0, sticky="w")
        self.var_label = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.var_label, width=30).grid(row=0, column=1, sticky="w")

        ttk.Label(frm_top, text="엑셀 파일:").grid(row=1, column=0, sticky="w")
        self.var_file = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.var_file, width=48, state="readonly").grid(
            row=1, column=1, columnspan=2, sticky="we")
        ttk.Button(frm_top, text="파일 선택", command=self._choose_file).grid(row=1, column=3, padx=4)

        ttk.Label(frm_top, text="시트:").grid(row=2, column=0, sticky="w")
        self.var_sheet = tk.StringVar()
        self.cb_sheet = ttk.Combobox(frm_top, textvariable=self.var_sheet, state="readonly", width=30)
        self.cb_sheet.grid(row=2, column=1, sticky="w")
        self.cb_sheet.bind("<<ComboboxSelected>>", lambda e: self._load_headers())

        ttk.Label(frm_top, text="검색할 열:").grid(row=3, column=0, sticky="w")
        self.var_search_col = tk.StringVar()
        self.cb_search_col = ttk.Combobox(frm_top, textvariable=self.var_search_col, state="readonly", width=30)
        self.cb_search_col.grid(row=3, column=1, sticky="w")

        ttk.Label(frm_top, text="검색 방식:").grid(row=4, column=0, sticky="w")
        self.var_match = tk.StringVar(value="contains")
        frm_match = ttk.Frame(frm_top)
        frm_match.grid(row=4, column=1, sticky="w")
        ttk.Radiobutton(frm_match, text="부분일치(포함)", variable=self.var_match, value="contains").pack(side="left")
        ttk.Radiobutton(frm_match, text="완전일치", variable=self.var_match, value="exact").pack(side="left")

        # 가져올 열 매핑 영역
        frm_cols = ttk.LabelFrame(self, text="가져올 열 (원본 열 -> 출력 별칭)")
        frm_cols.pack(fill="both", expand=True, **pad)

        self.tree_cols = ttk.Treeview(frm_cols, columns=("source", "alias"), show="headings", height=10)
        self.tree_cols.heading("source", text="원본 열(헤더명)")
        self.tree_cols.heading("alias", text="출력 별칭")
        self.tree_cols.column("source", width=250)
        self.tree_cols.column("alias", width=250)
        self.tree_cols.pack(fill="both", expand=True, padx=6, pady=4)

        frm_add = ttk.Frame(frm_cols)
        frm_add.pack(fill="x", padx=6, pady=4)

        ttk.Label(frm_add, text="원본 열:").pack(side="left")
        self.var_new_source = tk.StringVar()
        self.cb_new_source = ttk.Combobox(frm_add, textvariable=self.var_new_source, state="readonly", width=22)
        self.cb_new_source.pack(side="left", padx=4)
        self.cb_new_source.bind("<<ComboboxSelected>>", self._sync_alias_default)

        ttk.Label(frm_add, text="별칭:").pack(side="left")
        self.var_new_alias = tk.StringVar()
        ttk.Entry(frm_add, textvariable=self.var_new_alias, width=20).pack(side="left", padx=4)

        ttk.Button(frm_add, text="목록에 추가", command=self._add_column_map).pack(side="left", padx=6)
        ttk.Button(frm_add, text="선택 삭제", command=self._remove_column_map).pack(side="left")

        # 하단 저장/취소
        frm_bottom = ttk.Frame(self)
        frm_bottom.pack(fill="x", **pad)
        ttk.Button(frm_bottom, text="저장", command=self._on_save).pack(side="right", padx=4)
        ttk.Button(frm_bottom, text="취소", command=self.destroy).pack(side="right")

    # -- 동작 --------------------------------------------------------
    def _choose_file(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        self.var_file.set(path)
        try:
            sheets = SheetCache.list_sheets(path)
        except Exception as e:
            messagebox.showerror("오류", f"파일을 여는 중 오류가 발생했습니다:\n{e}", parent=self)
            return
        self.cb_sheet["values"] = sheets
        if sheets:
            self.var_sheet.set(sheets[0])
            self._load_headers()

    def _load_headers(self):
        path = self.var_file.get()
        sheet = self.var_sheet.get()
        if not path or not sheet:
            return
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb[sheet]
            first_row = next(ws.iter_rows(values_only=True), ())
            wb.close()
        except Exception as e:
            messagebox.showerror("오류", f"시트를 읽는 중 오류가 발생했습니다:\n{e}", parent=self)
            return

        headers = [("" if h is None else str(h).strip()) for h in first_row]
        headers = [h for h in headers if h != ""]
        self.available_headers = headers

        self.cb_search_col["values"] = headers
        self.cb_new_source["values"] = headers
        if headers and not self.var_search_col.get():
            self.var_search_col.set(headers[0])

    def _sync_alias_default(self, event=None):
        if not self.var_new_alias.get():
            self.var_new_alias.set(self.var_new_source.get())

    def _add_column_map(self):
        src = self.var_new_source.get().strip()
        alias = self.var_new_alias.get().strip()
        if not src:
            messagebox.showwarning("확인", "원본 열을 선택하세요.", parent=self)
            return
        if not alias:
            alias = src
        self.tree_cols.insert("", "end", values=(src, alias))
        self.var_new_source.set("")
        self.var_new_alias.set("")

    def _remove_column_map(self):
        for item in self.tree_cols.selection():
            self.tree_cols.delete(item)

    def _load_from_config(self, config: SourceConfig):
        self.var_label.set(config.label)
        self.var_file.set(config.file_path)
        try:
            sheets = SheetCache.list_sheets(config.file_path)
            self.cb_sheet["values"] = sheets
        except Exception:
            pass
        self.var_sheet.set(config.sheet_name)
        self._load_headers()
        self.var_search_col.set(config.search_header)
        self.var_match.set(config.match_mode)
        for cm in config.columns:
            self.tree_cols.insert("", "end", values=(cm.source_header, cm.alias))

    def _on_save(self):
        label = self.var_label.get().strip()
        file_path = self.var_file.get().strip()
        sheet = self.var_sheet.get().strip()
        search_col = self.var_search_col.get().strip()

        if not label:
            messagebox.showwarning("확인", "탭 이름을 입력하세요.", parent=self)
            return
        if not file_path or not os.path.isfile(file_path):
            messagebox.showwarning("확인", "유효한 엑셀 파일을 선택하세요.", parent=self)
            return
        if not sheet:
            messagebox.showwarning("확인", "시트를 선택하세요.", parent=self)
            return
        if not search_col:
            messagebox.showwarning("확인", "검색할 열을 선택하세요.", parent=self)
            return

        columns = []
        for item in self.tree_cols.get_children():
            src, alias = self.tree_cols.item(item, "values")
            columns.append(ColumnMap(source_header=src, alias=alias))
        if not columns:
            messagebox.showwarning("확인", "가져올 열을 최소 1개 이상 추가하세요.", parent=self)
            return

        enabled = self._editing.enabled if self._editing is not None else True

        self.result = SourceConfig(
            label=label,
            file_path=file_path,
            sheet_name=sheet,
            search_header=search_col,
            match_mode=self.var_match.get(),
            enabled=enabled,
            columns=columns,
        )
        self.destroy()


# ---------------------------------------------------------------------------
# 메인 애플리케이션
# ---------------------------------------------------------------------------

class ExcelSearchMergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("엑셀 검색/취합 도구")
        self.geometry("900x680")

        self.configs: List[SourceConfig] = []
        self.output_columns: List[OutputColumn] = []
        self.cache = SheetCache()
        self.last_results: List[Dict[str, object]] = []

        self._build_ui()
        self._load_config(silent=True)  # 이전에 저장한 상태가 있으면 자동으로 불러오기

    # ------------------------------------------------------------------
    def _build_ui(self):
        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=6, pady=(6, 0))

        ttk.Button(toolbar, text="상태 저장", command=self._save_config).pack(side="left", padx=2)
        ttk.Button(toolbar, text="상태 불러오기", command=lambda: self._load_config(silent=False)).pack(
            side="left", padx=2)
        ttk.Button(toolbar, text="초기화", command=self._reset_all).pack(side="left", padx=2)
        ttk.Label(toolbar, text=f"저장 위치: {DEFAULT_CONFIG_PATH}", foreground="#666").pack(side="left", padx=10)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=6, pady=6)

        self.tab_sources = ttk.Frame(nb)
        self.tab_output = ttk.Frame(nb)
        self.tab_run = ttk.Frame(nb)

        nb.add(self.tab_sources, text="① 소스(탭) 설정")
        nb.add(self.tab_output, text="② 출력 열 설정")
        nb.add(self.tab_run, text="③ 검색 실행 / 결과")

        self._build_tab_sources()
        self._build_tab_output()
        self._build_tab_run()

    # -- 탭 1: 소스 설정 --------------------------------------------------
    def _build_tab_sources(self):
        frm = self.tab_sources

        self.tree_sources = ttk.Treeview(
            frm, columns=("use", "label", "file", "sheet", "search", "cols"),
            show="headings", height=14
        )
        headers = [
            ("use", "사용", 50), ("label", "탭이름", 120), ("file", "파일", 220),
            ("sheet", "시트", 100), ("search", "검색열", 100), ("cols", "가져올열 수", 90),
        ]
        for key, text, w in headers:
            self.tree_sources.heading(key, text=text)
            self.tree_sources.column(key, width=w, anchor="center" if key in ("use", "cols") else "w")
        self.tree_sources.pack(fill="both", expand=True, padx=8, pady=8)

        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill="x", padx=8, pady=4)
        ttk.Button(btn_frm, text="추가", command=self._add_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="편집", command=self._edit_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="삭제", command=self._delete_source).pack(side="left", padx=2)
        ttk.Button(btn_frm, text="사용/제외 전환", command=self._toggle_source).pack(side="left", padx=2)

    def _refresh_sources_tree(self):
        self.tree_sources.delete(*self.tree_sources.get_children())
        for i, c in enumerate(self.configs):
            self.tree_sources.insert(
                "", "end", iid=str(i),
                values=("O" if c.enabled else "X", c.label, os.path.basename(c.file_path),
                        c.sheet_name, c.search_header, len(c.columns))
            )
        self._refresh_output_field_choices()

    def _add_source(self):
        dlg = SourceConfigDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self.configs.append(dlg.result)
            self._refresh_sources_tree()

    def _edit_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            messagebox.showinfo("확인", "편집할 항목을 선택하세요.")
            return
        idx = int(sel[0])
        dlg = SourceConfigDialog(self, config=self.configs[idx])
        self.wait_window(dlg)
        if dlg.result:
            self.configs[idx] = dlg.result
            self._refresh_sources_tree()

    def _delete_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            return
        idx = int(sel[0])
        del self.configs[idx]
        self._refresh_sources_tree()

    def _toggle_source(self):
        sel = self.tree_sources.selection()
        if not sel:
            return
        idx = int(sel[0])
        self.configs[idx].enabled = not self.configs[idx].enabled
        self._refresh_sources_tree()

    # -- 탭 2: 출력 열 설정 -------------------------------------------------
    def _build_tab_output(self):
        frm = self.tab_output

        self.tree_output = ttk.Treeview(frm, columns=("header", "field"), show="headings", height=12)
        self.tree_output.heading("header", text="출력 헤더명")
        self.tree_output.heading("field", text="필드(소스)")
        self.tree_output.column("header", width=250)
        self.tree_output.column("field", width=250)
        self.tree_output.pack(fill="both", expand=True, padx=8, pady=8)

        frm_add = ttk.Frame(frm)
        frm_add.pack(fill="x", padx=8, pady=4)

        ttk.Label(frm_add, text="필드:").pack(side="left")
        self.var_out_field = tk.StringVar()
        self.cb_out_field = ttk.Combobox(frm_add, textvariable=self.var_out_field, state="readonly", width=28)
        self.cb_out_field.pack(side="left", padx=4)

        ttk.Label(frm_add, text="헤더명:").pack(side="left")
        self.var_out_header = tk.StringVar()
        ttk.Entry(frm_add, textvariable=self.var_out_header, width=22).pack(side="left", padx=4)

        ttk.Button(frm_add, text="추가", command=self._add_output_column).pack(side="left", padx=6)
        ttk.Button(frm_add, text="선택 삭제", command=self._remove_output_column).pack(side="left", padx=2)
        ttk.Button(frm_add, text="위로", command=lambda: self._move_output_column(-1)).pack(side="left", padx=2)
        ttk.Button(frm_add, text="아래로", command=lambda: self._move_output_column(1)).pack(side="left", padx=2)

        ttk.Label(
            frm,
            text="※ 필드는 [검색어], [출처 탭이름], 또는 소스 설정에서 지정한 '출력 별칭' 중에서 고를 수 있습니다.",
            foreground="#555"
        ).pack(anchor="w", padx=8, pady=(0, 8))

    def _refresh_output_field_choices(self):
        aliases = []
        for c in self.configs:
            for cm in c.columns:
                if cm.alias not in aliases:
                    aliases.append(cm.alias)
        choices = list(SPECIAL_FIELDS.values()) + aliases
        self.cb_out_field["values"] = choices

    def _field_label_to_key(self, label: str) -> str:
        for k, v in SPECIAL_FIELDS.items():
            if v == label:
                return k
        return label  # alias 그대로 키로 사용

    def _field_key_to_label(self, key: str) -> str:
        return SPECIAL_FIELDS.get(key, key)

    def _refresh_output_tree(self):
        self.tree_output.delete(*self.tree_output.get_children())
        for oc in self.output_columns:
            self.tree_output.insert("", "end", values=(oc.header, self._field_key_to_label(oc.field_key)))

    def _add_output_column(self):
        field_label = self.var_out_field.get().strip()
        header = self.var_out_header.get().strip()
        if not field_label:
            messagebox.showwarning("확인", "필드를 선택하세요.", parent=self)
            return
        if not header:
            header = field_label
        key = self._field_label_to_key(field_label)
        self.output_columns.append(OutputColumn(field_key=key, header=header))
        self._refresh_output_tree()
        self.var_out_field.set("")
        self.var_out_header.set("")

    def _remove_output_column(self):
        sel = self.tree_output.selection()
        if not sel:
            return
        indices = sorted((self.tree_output.index(i) for i in sel), reverse=True)
        for i in indices:
            del self.output_columns[i]
        self._refresh_output_tree()

    def _move_output_column(self, delta: int):
        sel = self.tree_output.selection()
        if not sel:
            return
        idx = self.tree_output.index(sel[0])
        new_idx = idx + delta
        if 0 <= new_idx < len(self.output_columns):
            self.output_columns[idx], self.output_columns[new_idx] = (
                self.output_columns[new_idx], self.output_columns[idx]
            )
            self._refresh_output_tree()
            self.tree_output.selection_set(str(new_idx))

    # -- 탭 3: 검색 실행 / 결과 ---------------------------------------------
    def _build_tab_run(self):
        frm = self.tab_run

        ttk.Label(frm, text="검색어 입력 (한 줄에 하나씩):").pack(anchor="w", padx=8, pady=(8, 0))
        self.txt_terms = tk.Text(frm, height=6)
        self.txt_terms.pack(fill="x", padx=8, pady=4)

        opt_frm = ttk.Frame(frm)
        opt_frm.pack(fill="x", padx=8)
        self.var_case_insensitive = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_frm, text="대소문자 구분 안 함", variable=self.var_case_insensitive).pack(side="left")

        self.var_merge_mode = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frm, text="여러 탭 결과를 검색어 기준으로 한 행에 병합(권장)",
            variable=self.var_merge_mode
        ).pack(side="left", padx=16)

        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill="x", padx=8, pady=6)
        ttk.Button(btn_frm, text="검색 실행", command=self._run_search).pack(side="left")
        ttk.Button(btn_frm, text="엑셀로 저장", command=self._save_results).pack(side="left", padx=6)

        self.lbl_status = ttk.Label(frm, text="검색 결과: 0건")
        self.lbl_status.pack(anchor="w", padx=8)

        self.tree_result = ttk.Treeview(frm, show="headings", height=16)
        self.tree_result.pack(fill="both", expand=True, padx=8, pady=8)

    def _run_search(self):
        if not self.configs:
            messagebox.showwarning("확인", "소스(탭) 설정을 먼저 추가하세요.")
            return
        if not self.output_columns:
            messagebox.showwarning("확인", "출력 열 설정을 먼저 추가하세요.")
            return

        terms = [t.strip() for t in self.txt_terms.get("1.0", "end").splitlines() if t.strip()]
        if not terms:
            messagebox.showwarning("확인", "검색어를 한 개 이상 입력하세요.")
            return

        self.cache.clear()
        results: List[Dict[str, object]] = []
        case_ins = self.var_case_insensitive.get()
        merge_mode = self.var_merge_mode.get()

        try:
            active_configs = [c for c in self.configs if c.enabled]
            for term in terms:
                if merge_mode:
                    # 병합 모드: 검색어 하나당, 활성화된 탭들의 매치 결과를 한 데이터셋으로 묶는다.
                    # - 어떤 탭에서 매치가 여러 건이면(예: 구매내역) 그 건수만큼 행을 만든다.
                    # - 매치가 1건뿐인 탭(예: 직원목록)의 값은 모든 행에 동일하게 채워 넣는다(반복).
                    per_config_matches = []
                    for cfg in active_configs:
                        headers, rows = self.cache.get(cfg.file_path, cfg.sheet_name)
                        matches = search_in_config(cfg, headers, rows, term, case_ins)
                        per_config_matches.append((cfg, matches))

                    match_counts = [len(m) for _, m in per_config_matches if len(m) > 0]
                    if not match_counts:
                        continue  # 이 검색어는 어떤 탭에서도 매치되지 않음
                    max_count = max(match_counts)

                    for i in range(max_count):
                        rec: Dict[str, object] = {SEARCH_TERM_KEY: term}
                        labels = []
                        for cfg, matches in per_config_matches:
                            if not matches:
                                continue
                            labels.append(cfg.label)
                            if len(matches) == 1:
                                rec.update(matches[0])       # 1건이면 모든 행에 반복 채움
                            elif i < len(matches):
                                rec.update(matches[i])       # 여러 건이면 순서대로 배치
                            # i가 해당 탭의 매치 건수를 넘어서면 그 탭의 값은 빈칸으로 둠
                        rec[SOURCE_LABEL_KEY] = "+".join(labels)
                        results.append(rec)
                else:
                    # 쌓기 모드(기존 방식): 탭별로 매치된 행을 각각 별도의 결과 행으로 추가
                    for cfg in active_configs:
                        headers, rows = self.cache.get(cfg.file_path, cfg.sheet_name)
                        matches = search_in_config(cfg, headers, rows, term, case_ins)
                        for m in matches:
                            rec = dict(m)
                            rec[SEARCH_TERM_KEY] = term
                            rec[SOURCE_LABEL_KEY] = cfg.label
                            results.append(rec)
        except Exception as e:
            messagebox.showerror("오류", f"검색 중 오류가 발생했습니다:\n{e}\n\n{traceback.format_exc()}")
            return

        self.last_results = results
        self._show_results(results)

    def _show_results(self, results: List[Dict[str, object]]):
        cols = [oc.header for oc in self.output_columns]
        self.tree_result["columns"] = cols
        self.tree_result.delete(*self.tree_result.get_children())
        for c in cols:
            self.tree_result.heading(c, text=c)
            self.tree_result.column(c, width=140, anchor="w")

        for rec in results:
            row_vals = []
            for oc in self.output_columns:
                v = rec.get(oc.field_key, "")
                row_vals.append("" if v is None else v)
            self.tree_result.insert("", "end", values=row_vals)

        self.lbl_status.config(text=f"검색 결과: {len(results)}건")

    def _save_results(self):
        if not self.last_results:
            messagebox.showinfo("확인", "먼저 검색을 실행하세요.")
            return

        path = filedialog.asksaveasfilename(
            title="결과 저장", defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")]
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "결과"

            headers = [oc.header for oc in self.output_columns]
            ws.append(headers)

            for rec in self.last_results:
                row = []
                for oc in self.output_columns:
                    v = rec.get(oc.field_key, "")
                    row.append("" if v is None else v)
                ws.append(row)

            # 대략적인 열 너비 자동 조정
            for col_idx, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for rec in self.last_results:
                    v = rec.get(self.output_columns[col_idx - 1].field_key, "")
                    max_len = max(max_len, len(str(v)))
                letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[letter].width = min(max_len + 4, 60)

            wb.save(path)
        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류가 발생했습니다:\n{e}")
            return

        messagebox.showinfo("완료", f"저장이 완료되었습니다:\n{path}")
        open_file_with_default_app(path)

    # -- 상태 저장 / 불러오기 / 초기화 -------------------------------------
    def _config_to_dict(self) -> dict:
        return {
            "sources": [
                {
                    "label": c.label,
                    "file_path": c.file_path,
                    "sheet_name": c.sheet_name,
                    "search_header": c.search_header,
                    "match_mode": c.match_mode,
                    "enabled": c.enabled,
                    "columns": [{"source_header": cm.source_header, "alias": cm.alias} for cm in c.columns],
                }
                for c in self.configs
            ],
            "output_columns": [
                {"field_key": oc.field_key, "header": oc.header} for oc in self.output_columns
            ],
        }

    def _apply_config_dict(self, data: dict):
        configs = []
        for cd in data.get("sources", []):
            cols = [
                ColumnMap(source_header=cm.get("source_header", ""), alias=cm.get("alias", ""))
                for cm in cd.get("columns", [])
            ]
            configs.append(SourceConfig(
                label=cd.get("label", ""),
                file_path=cd.get("file_path", ""),
                sheet_name=cd.get("sheet_name", ""),
                search_header=cd.get("search_header", ""),
                match_mode=cd.get("match_mode", "contains"),
                enabled=cd.get("enabled", True),
                columns=cols,
            ))
        output_columns = [
            OutputColumn(field_key=ocd.get("field_key", ""), header=ocd.get("header", ""))
            for ocd in data.get("output_columns", [])
        ]

        self.configs = configs
        self.output_columns = output_columns
        self._refresh_sources_tree()
        self._refresh_output_tree()

    def _save_config(self):
        data = self._config_to_dict()
        try:
            with open(DEFAULT_CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("오류", f"상태 저장 중 오류가 발생했습니다:\n{e}")
            return
        messagebox.showinfo("완료", f"현재 소스/출력 설정을 저장했습니다:\n{DEFAULT_CONFIG_PATH}")

    def _load_config(self, silent: bool = False):
        if not os.path.isfile(DEFAULT_CONFIG_PATH):
            if not silent:
                messagebox.showinfo("확인", "저장된 상태 파일이 없습니다.")
            return
        try:
            with open(DEFAULT_CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            self._apply_config_dict(data)
        except Exception as e:
            if not silent:
                messagebox.showerror("오류", f"상태 불러오기 중 오류가 발생했습니다:\n{e}")
            return
        if not silent:
            messagebox.showinfo("완료", "저장된 소스/출력 설정을 불러왔습니다.")

    def _reset_all(self):
        if not messagebox.askyesno("확인", "소스 설정, 출력 열 설정, 검색어, 결과를 모두 초기화하시겠습니까?\n"
                                          "(저장된 상태 파일은 삭제되지 않습니다.)"):
            return
        self.configs = []
        self.output_columns = []
        self.last_results = []
        self._refresh_sources_tree()
        self._refresh_output_tree()
        self.txt_terms.delete("1.0", "end")
        self.tree_result["columns"] = ()
        self.tree_result.delete(*self.tree_result.get_children())
        self.lbl_status.config(text="검색 결과: 0건")


def main():
    app = ExcelSearchMergerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
